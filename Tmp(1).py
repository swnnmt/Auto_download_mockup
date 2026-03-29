import os
import re
import time
from pathlib import Path
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse, parse_qsl

import traceback
import requests
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException

BASE_DIR = Path(__file__).resolve().parent
LINK_XLSX = BASE_DIR / "LinkMockup.xlsx"
BATH_IMAGE_XLSX = BASE_DIR / "Bath_image.xlsx"
IMAGE_SRC_DIR = BASE_DIR / "Image_src"

SHEET_NAME = "Bath_Image"
SLEEP_AFTER_CROP = 20
PAGE_WAIT = 20
MAX_LINKS = None
HEADLESS = False

USER_DATA_DIR = Path(os.environ.get("CHROME_USER_DATA_DIR", str(BASE_DIR / "ChromeProfileBot")))
PROFILE_DIRECTORY = os.environ.get("CHROME_PROFILE_DIR", "")
RUN_MODE = os.environ.get("RUN_MODE", "selected").strip().lower()
VALID_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def read_mockup_links(xlsx_path: Path, max_links: int | None = None):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    links = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        raw = normalize_text(row[0])
        if raw:
            links.append(raw)

        if max_links is not None and len(links) >= max_links:
            break

    return links


def load_batch_sheet(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]
    headers = {}
    for cell in ws[1]:
        headers[normalize_text(cell.value)] = cell.column

    required = ["Theme", "ImageName", "ImagePath", "Selected", "Status", "Note"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError(f"Bath_image.xlsx thiếu cột: {', '.join(missing)}")

    return wb, ws, headers


def should_process(selected: str, status: str):
    selected = selected.upper()
    status = status.capitalize()

    if RUN_MODE == "selected":
        return selected in {"Y", "R"}
    if RUN_MODE == "pending":
        if status == "Done":
            return False
        return status in {"", "Pending", "Error"} or selected in {"Y", "R"}
    raise RuntimeError("RUN_MODE chỉ hỗ trợ: selected hoặc pending")


def update_row(ws, row_idx, headers, status=None, note=None):
    if status is not None:
        ws.cell(row=row_idx, column=headers["Status"]).value = status
    if note is not None:
        ws.cell(row=row_idx, column=headers["Note"]).value = note


def find_chrome_binary() -> Path:
    env_binary = normalize_text(os.environ.get("CHROME_BINARY"))
    if env_binary and Path(env_binary).exists():
        return Path(env_binary)

    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path.home() / "AppData" / "Local" / "Google" / "Chrome" / "Application" / "chrome.exe",
        Path(r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"),
        Path(r"C:\Program Files (x86)\BraveSoftware\Brave-Browser\Application\brave.exe"),
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(
        "Không tìm thấy chrome.exe. Hãy đặt biến môi trường CHROME_BINARY tới file chrome.exe."
    )


def find_chromedriver() -> Path | None:
    env_driver = normalize_text(os.environ.get("CHROMEDRIVER"))
    if env_driver and Path(env_driver).exists():
        return Path(env_driver)

    candidates = [
        BASE_DIR / "chromedriver.exe",
        Path.cwd() / "chromedriver.exe",
        Path(r"C:\chromedriver\chromedriver.exe"),
    ]
    for path in candidates:
        if path.exists():
            return path
    return None


def build_driver():
    if not USER_DATA_DIR.exists():
        raise FileNotFoundError(f"Không thấy Chrome user data dir: {USER_DATA_DIR}")

    chrome_binary = find_chrome_binary()

    options = webdriver.ChromeOptions()
    options.binary_location = str(chrome_binary)

    if HEADLESS:
        options.add_argument("--headless=new")

    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"--user-data-dir={USER_DATA_DIR}")
    if PROFILE_DIRECTORY:
        options.add_argument(f"--profile-directory={PROFILE_DIRECTORY}")
    options.add_argument("--remote-allow-origins=*")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    prefs = {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)

    chromedriver_path = find_chromedriver()
    if chromedriver_path:
        print(f"Dùng chromedriver local: {chromedriver_path}")
        service = Service(executable_path=str(chromedriver_path))
        driver = webdriver.Chrome(service=service, options=options)
    else:
        print("Không thấy chromedriver local, thử dùng Selenium Manager...")
        driver = webdriver.Chrome(options=options)

    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            """
        },
    )
    return driver


def try_click(wait, selectors):
    last_exc = None
    for by, value in selectors:
        try:
            el = wait.until(EC.element_to_be_clickable((by, value)))
            el.click()
            return
        except Exception as exc:
            last_exc = exc
    if last_exc:
        raise last_exc


def close_popups_if_any(driver):
    possible = [
        (By.XPATH, "//button[contains(@aria-label, 'Close')]"),
        (By.XPATH, "//button[contains(., 'Got it')]"),
        (By.XPATH, "//button[contains(., 'No thanks')]"),
        (By.XPATH, "//button[contains(., 'Maybe later')]"),
    ]
    for by, value in possible:
        try:
            buttons = driver.find_elements(by, value)
            for btn in buttons[:2]:
                try:
                    btn.click()
                    time.sleep(1)
                except Exception:
                    pass
        except Exception:
            pass


def click_continue_as_if_present(driver, timeout=25):
    end_time = time.time() + timeout

    def try_click(elem, label="elem"):
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center', inline:'center'});",
                elem
            )
        except Exception:
            pass

        time.sleep(0.2)

        for method in ("normal", "action", "js"):
            try:
                if method == "normal":
                    elem.click()
                elif method == "action":
                    ActionChains(driver).move_to_element(elem).pause(0.2).click().perform()
                else:
                    driver.execute_script("arguments[0].click();", elem)

                print(f"[DEBUG] Đã click Continue as bằng {method}: {label}")
                return True
            except Exception as e:
                print(f"[DEBUG] Click {method} fail {label}: {type(e).__name__}: {e}")

        return False

    candidate_xpaths = [
        "//button[contains(., 'Continue as')]",
        "//div[@role='button' and contains(., 'Continue as')]",
        "//*[contains(@aria-label, 'Continue as')]",
        "//*[contains(normalize-space(.), 'Continue as')]",
        "//*[contains(normalize-space(.), 'Continue as Thắng')]",
    ]

    while time.time() < end_time:
        driver.switch_to.default_content()

        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        print(f"[DEBUG] Số iframe hiện có: {len(iframes)}")

        for xp in candidate_xpaths:
            try:
                elems = driver.find_elements(By.XPATH, xp)
                print(f"[DEBUG] Main xpath={xp} -> {len(elems)} elements")
                for i, elem in enumerate(elems):
                    try:
                        txt = (elem.text or "").strip()
                    except Exception:
                        txt = ""
                    print(f"[DEBUG] Main elem {i} text={txt!r}")
                    if elem.is_displayed() and elem.is_enabled():
                        if try_click(elem, f"main/{xp}/{i}"):
                            driver.switch_to.default_content()
                            time.sleep(3)
                            return True
            except Exception as e:
                print(f"[DEBUG] Main xpath lỗi {xp}: {type(e).__name__}: {e}")

        iframe_infos = []
        for idx, frame in enumerate(iframes):
            try:
                src = frame.get_attribute("src") or ""
                title = frame.get_attribute("title") or ""
                name = frame.get_attribute("name") or ""
                iframe_infos.append((idx, frame, src, title, name))
                print(f"[DEBUG] iframe[{idx}] src={src!r} title={title!r} name={name!r}")
            except Exception:
                pass

        iframe_infos.sort(
            key=lambda item: 0 if ("google" in (item[2] or "").lower() or "gsi" in (item[2] or "").lower() or "google" in (item[3] or "").lower()) else 1
        )

        for idx, frame, src, title, name in iframe_infos:
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                print(f"[DEBUG] Đang kiểm tra iframe[{idx}]")

                for xp in candidate_xpaths:
                    elems = driver.find_elements(By.XPATH, xp)
                    print(f"[DEBUG] iframe[{idx}] xpath={xp} -> {len(elems)} elements")
                    for j, elem in enumerate(elems):
                        try:
                            txt = (elem.text or "").strip()
                        except Exception:
                            txt = ""
                        print(f"[DEBUG] iframe[{idx}] elem {j} text={txt!r}")

                        try:
                            if elem.is_displayed() and elem.is_enabled():
                                if try_click(elem, f"iframe[{idx}]/{xp}/{j}"):
                                    driver.switch_to.default_content()
                                    time.sleep(3)
                                    return True
                        except Exception:
                            pass

                css_candidates = [
                    "button",
                    "[role='button']",
                    "div[role='button']",
                ]
                for css in css_candidates:
                    elems = driver.find_elements(By.CSS_SELECTOR, css)
                    for j, elem in enumerate(elems):
                        try:
                            txt = (elem.text or "").strip()
                            if "Continue as" in txt and elem.is_displayed() and elem.is_enabled():
                                if try_click(elem, f"iframe[{idx}]/{css}/{j}"):
                                    driver.switch_to.default_content()
                                    time.sleep(3)
                                    return True
                        except Exception:
                            pass

            except Exception as e:
                print(f"[DEBUG] Không vào được iframe[{idx}]: {type(e).__name__}: {e}")

        driver.switch_to.default_content()
        time.sleep(1)

    print("[DEBUG] Không thấy popup Continue as")
    return False


def ensure_placeit_login(driver, first_link: str):
    print("[DEBUG] Bắt đầu login đầu phiên Placeit...")
    driver.get(first_link)

    wait = WebDriverWait(driver, PAGE_WAIT)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(5)

    clicked = click_continue_as_if_present(driver, timeout=25)
    if clicked:
        print("[DEBUG] Đã login Placeit bằng Continue as")
        time.sleep(3)
    else:
        print("[DEBUG] Không click được Continue as")

    close_popups_if_any(driver)
    time.sleep(1)


def click_crop(driver, timeout=20):
    wait = WebDriverWait(driver, timeout)

    crop_btn = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "button.cropButton"))
    )

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", crop_btn)

    wait.until(lambda d: crop_btn.is_displayed() and crop_btn.is_enabled())

    try:
        crop_btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", crop_btn)


def upload_design_and_get_custom_param(driver, first_link: str, image_path: Path) -> str:
    wait = WebDriverWait(driver, PAGE_WAIT)
    driver.get(first_link)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(1)
    close_popups_if_any(driver)

    try:
        try_click(wait, [
            (By.XPATH, "//button[contains(., 'Design')]"),
            (By.CSS_SELECTOR, ".custom-graphic-dropdown button.dropdown-toggle"),
            (By.CSS_SELECTOR, "button.btn.dropdown-toggle.btn-default"),
        ])
    except Exception:
        pass

    drop_zone = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, ".drop-zone[data-drop-target-id]")
        )
    )

    target_input_id = drop_zone.get_attribute("data-drop-target-id")
    if not target_input_id:
        raise RuntimeError("Không lấy được data-drop-target-id của vùng upload.")

    print(f"[DEBUG] target input id = {target_input_id}")

    file_input = wait.until(
        EC.presence_of_element_located((By.ID, target_input_id))
    )

    file_input.send_keys(str(image_path.resolve()))
    print(f"[DEBUG] Đã gửi file: {image_path.resolve()}")

    try:
        WebDriverWait(driver, 20).until(
            lambda d: (
                "customG_0=" in d.current_url
                or len(d.find_elements(By.XPATH, "//button[contains(., 'Crop') or contains(., 'Apply') or contains(., 'Done')]")) > 0
            )
        )
    except Exception:
        pass

    try:
        click_crop(driver, timeout=20)
        print("[DEBUG] Đã click Crop")
    except Exception as exc:
        print(f"[DEBUG] Không click được Crop: {exc}")

    WebDriverWait(driver, 30).until(
        lambda d: "customG_0=" in d.current_url
    )

    current_url = driver.current_url
    print("[DEBUG] URL sau upload:", current_url)

    query = parse_qs(urlparse(current_url).query)
    custom_value = query.get("customG_0", [None])[0]
    if custom_value:
        return f"customG_0={custom_value}"

    m = re.search(r"[?&](customG_0=[^&]+)", current_url)
    if m:
        return m.group(1)

    raise RuntimeError(f"Không lấy được customG_0 từ URL: {current_url}")


def append_or_replace_custom_param(url: str, custom_param: str) -> str:
    key, value = custom_param.split("=", 1)
    parsed = urlparse(url)
    params = dict(parse_qsl(parsed.query, keep_blank_values=True))
    params[key] = value
    new_query = urlencode(params)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def click_element_robust(driver, element):
    try:
        element.click()
        print("[DEBUG] Click thường thành công")
        return True
    except Exception as e:
        print(f"[DEBUG] Click thường thất bại: {type(e).__name__}: {e}")

    try:
        ActionChains(driver).move_to_element(element).pause(0.5).click().perform()
        print("[DEBUG] ActionChains click thành công")
        return True
    except Exception as e:
        print(f"[DEBUG] ActionChains click thất bại: {type(e).__name__}: {e}")

    try:
        driver.execute_script("arguments[0].click();", element)
        print("[DEBUG] JS click thành công")
        return True
    except Exception as e:
        print(f"[DEBUG] JS click thất bại: {type(e).__name__}: {e}")

    return False


def wait_for_any_download_button(driver, timeout=30):
    wait = WebDriverWait(driver, timeout)

    js = """
    const buttons = Array.from(document.querySelectorAll('button.download-button'));
    const visible = buttons.filter(btn => {
        const rect = btn.getBoundingClientRect();
        const text = (btn.innerText || btn.textContent || '').trim();
        const style = window.getComputedStyle(btn);
        return text === 'Download'
            && rect.width > 0
            && rect.height > 0
            && style.display !== 'none'
            && style.visibility !== 'hidden'
            && rect.top >= 0
            && rect.left >= 0;
    });

    if (!visible.length) return null;

    visible.sort((a, b) => {
        const ra = a.getBoundingClientRect();
        const rb = b.getBoundingClientRect();
        const scoreA = ra.top * 1000 - ra.left;
        const scoreB = rb.top * 1000 - rb.left;
        return scoreA - scoreB;
    });

    return visible[0];
    """

    btn = wait.until(lambda d: d.execute_script(js))
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", btn)
    time.sleep(0.8)
    return btn


def get_real_download_url(driver, page_url: str) -> str | None:
    print(f"[DEBUG] Mở trang mockup: {page_url}")
    driver.get(page_url)

    wait = WebDriverWait(driver, PAGE_WAIT)

    try:
        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(3)

        close_popups_if_any(driver)
        time.sleep(1)

        print(f"[DEBUG] URL hiện tại: {driver.current_url}")
        print(f"[DEBUG] Title: {driver.title}")

        old_links = set()
        try:
            old_link_elems = driver.find_elements(
                By.XPATH,
                "//a[contains(normalize-space(.), 'Click here to download')]"
            )
            for a in old_link_elems:
                href = (a.get_attribute("href") or "").strip()
                if href.startswith("http"):
                    old_links.add(href)
        except Exception:
            pass

        print(f"[DEBUG] Số link cũ trước khi click: {len(old_links)}")

        download_btn = wait_for_any_download_button(driver, timeout=PAGE_WAIT)
        print(f"[DEBUG] outerHTML nút đã chọn: {download_btn.get_attribute('outerHTML')}")

        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
                download_btn
            )
        except Exception:
            pass
        time.sleep(1)

        close_popups_if_any(driver)
        time.sleep(0.5)

        try:
            print(f"[DEBUG] is_displayed = {download_btn.is_displayed()}")
            print(f"[DEBUG] is_enabled = {download_btn.is_enabled()}")
            print(f"[DEBUG] class = {download_btn.get_attribute('class')}")
            print(f"[DEBUG] outerHTML = {download_btn.get_attribute('outerHTML')}")
        except Exception as e:
            print(f"[DEBUG] Không đọc được trạng thái nút: {type(e).__name__}: {e}")

        ok = click_element_robust(driver, download_btn)
        if not ok:
            raise RuntimeError("Không click được nút Download bằng mọi cách")

        print("[DEBUG] Đã click nút Download, chờ link mới xuất hiện...")

        def wait_for_new_download_link(driver, old_links, timeout=120):
            popup_wait = WebDriverWait(driver, timeout)

            popup_wait.until(
                lambda d: len(d.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'My Downloads')]")) > 0
            )

            def has_new_link(d):
                links = d.find_elements(
                    By.XPATH,
                    "//a[contains(normalize-space(.), 'Click here to download')]"
                )

                visible_processing = []
                try:
                    processing_nodes = d.find_elements(
                        By.XPATH,
                        "//*[contains(normalize-space(.), 'Processing your mockup')]"
                    )
                    visible_processing = [el for el in processing_nodes if el.is_displayed()]
                except Exception:
                    pass

                visible_links = 0
                for a in links:
                    try:
                        href = (a.get_attribute("href") or "").strip()
                        if a.is_displayed() and href.startswith("http"):
                            visible_links += 1

                            if href not in old_links:
                                print(f"[DEBUG] visible_processing={len(visible_processing)}, visible_links={visible_links}")
                                print(f"[DEBUG] Phát hiện link mới: {href}")
                                return href
                    except Exception:
                        pass

                print(f"[DEBUG] visible_processing={len(visible_processing)}, visible_links={visible_links}")
                return False

            return popup_wait.until(has_new_link)

        download_url = wait_for_new_download_link(driver, old_links, timeout=120)

        if download_url:
            print(f"[DEBUG] Real download URL mới: {download_url}")
            return download_url

        print("[DEBUG] Không lấy được link mới sau khi click Download")
        return None

    except Exception as exc:
        print(f"[ERROR] get_real_download_url thất bại: {type(exc).__name__}: {exc}")
        traceback.print_exc()
        return None


def copy_cookies_to_requests(driver, session: requests.Session):
    for cookie in driver.get_cookies():
        try:
            session.cookies.set(
                cookie["name"],
                cookie["value"],
                domain=cookie.get("domain"),
                path=cookie.get("path"),
            )
        except Exception:
            pass


def download_file(session: requests.Session, file_url: str, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with session.get(file_url, stream=True, timeout=120) as response:
        response.raise_for_status()
        with open(output_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=1024 * 128):
                if chunk:
                    f.write(chunk)


def update_mockup_status(folder_path: Path):
    status_file = BASE_DIR / "mockup_status.xlsx"
    if not status_file.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Mockup Status"
        ws.append(["bath_mockup", "json", "statusUpload", "status"])
        wb.save(status_file)

    wb = load_workbook(status_file)
    ws = wb.active
    
    # Kiểm tra và cập nhật header nếu là phiên bản cũ (2 hoặc 3 cột)
    if ws.max_column < 4:
        # Nếu có 3 cột (bath_mockup, json, status)
        if ws.max_column == 3:
            for row in range(2, ws.max_row + 1):
                old_status = ws.cell(row=row, column=3).value
                ws.cell(row=row, column=4).value = old_status # status sang cột 4
                ws.cell(row=row, column=3).value = "" # statusUpload trống ở cột 3
        # Nếu có 2 cột (bath_mockup, status)
        elif ws.max_column == 2:
            for row in range(2, ws.max_row + 1):
                old_status = ws.cell(row=row, column=2).value
                ws.cell(row=row, column=4).value = old_status
                ws.cell(row=row, column=2).value = ""
                ws.cell(row=row, column=3).value = ""
        
        # Cập nhật lại header
        ws.cell(row=1, column=1).value = "bath_mockup"
        ws.cell(row=1, column=2).value = "json"
        ws.cell(row=1, column=3).value = "statusUpload"
        ws.cell(row=1, column=4).value = "status"
        wb.save(status_file)

    # Check if the folder path already exists
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == str(folder_path):
            return

    ws.append([str(folder_path), "", "", "created"])
    wb.save(status_file)


def get_image_short_name(image_path: Path) -> str:
    name = image_path.stem.strip()
    if not name:
        return "Img"
    return name[:3]


def build_mockup_output_dir(image_path: Path) -> Path:
    short_name = get_image_short_name(image_path)
    return image_path.parent / f"IATE-{short_name}"


def build_mockup_output_file(image_path: Path, idx: int, total: int, ext: str) -> Path:
    short_name = get_image_short_name(image_path)
    prefix = f"IATE-{short_name}"

    split_point = (total + 1) // 2

    if idx <= split_point:
        group = "w"
        num = idx
    else:
        group = "b"
        num = idx - split_point

    output_dir = build_mockup_output_dir(image_path)
    filename = f"{prefix}-{group}-{num}{ext}"
    return output_dir / filename


def process_one_image(driver, session, links, image_path: Path):
    if not image_path.exists():
        raise FileNotFoundError(f"Không thấy file ảnh: {image_path}")
    if image_path.suffix.lower() not in VALID_IMAGE_EXTENSIONS:
        raise RuntimeError(f"File không phải ảnh hợp lệ: {image_path}")
    if not links:
        raise RuntimeError("Không có link mockup nào để xử lý.")

    output_dir = build_mockup_output_dir(image_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    update_mockup_status(output_dir)

    custom_param = upload_design_and_get_custom_param(driver, links[0], image_path)
    final_links = [append_or_replace_custom_param(link, custom_param) for link in links]

    total = len(final_links)
    saved_files = []

    for idx, link in enumerate(final_links, start=1):
        print(f"  -> Mockup {idx}/{total}")

        download_url = get_real_download_url(driver, link)
        print(f"     Download URL: {download_url}")

        if not download_url:
            raise RuntimeError(f"Không lấy được link download thật cho mockup {idx}")

        copy_cookies_to_requests(driver, session)

        parsed_path = urlparse(download_url).path or ""
        ext = Path(parsed_path).suffix.lower()
        if ext not in VALID_IMAGE_EXTENSIONS:
            ext = ".png"

        output_file = build_mockup_output_file(
            image_path=image_path,
            idx=idx,
            total=total,
            ext=ext,
        )

        download_file(session, download_url, output_file)
        saved_files.append(str(output_file))

    (output_dir / "generated_links.txt").write_text("\n".join(final_links), encoding="utf-8")
    (output_dir / "custom_param.txt").write_text(custom_param, encoding="utf-8")
    return custom_param, saved_files, output_dir


def main():
    if not LINK_XLSX.exists():
        raise FileNotFoundError(f"Không tìm thấy {LINK_XLSX}")
    if not BATH_IMAGE_XLSX.exists():
        raise FileNotFoundError(f"Không tìm thấy {BATH_IMAGE_XLSX}")
    if not IMAGE_SRC_DIR.exists():
        raise FileNotFoundError(f"Không tìm thấy thư mục {IMAGE_SRC_DIR}")

    links = read_mockup_links(LINK_XLSX, MAX_LINKS)
    if len(links) < 1:
        raise RuntimeError("LinkMockup.xlsx không có link nào.")
    print(f"Đọc được {len(links)} link mockup.")

    wb, ws, headers = load_batch_sheet(BATH_IMAGE_XLSX)
    driver = build_driver()
    session = requests.Session()

    try:
        # ensure_placeit_login(driver, links[0])
        for row_idx in range(2, ws.max_row + 1):
            theme = normalize_text(ws.cell(row=row_idx, column=headers["Theme"]).value)
            image_name = normalize_text(ws.cell(row=row_idx, column=headers["ImageName"]).value)
            image_path_raw = normalize_text(ws.cell(row=row_idx, column=headers["ImagePath"]).value)
            selected = normalize_text(ws.cell(row=row_idx, column=headers["Selected"]).value)
            status = normalize_text(ws.cell(row=row_idx, column=headers["Status"]).value)

            if not theme and not image_name and not image_path_raw:
                continue

            if not should_process(selected, status):
                continue

            try:
                print(f"\nĐang xử lý dòng {row_idx}: {theme} / {image_name}")
                update_row(ws, row_idx, headers, status="Running", note="")
                wb.save(BATH_IMAGE_XLSX)

                image_path = Path(image_path_raw)
                custom_param, saved_files, output_dir = process_one_image(
                    driver=driver,
                    session=session,
                    links=links,
                    image_path=image_path,
                )

                note = f"OK | {len(saved_files)} ảnh | {custom_param} | {output_dir}"
                update_row(ws, row_idx, headers, status="Done", note=note)

            except Exception as exc:
                update_row(ws, row_idx, headers, status="Error", note=str(exc)[:300])
                print(f"LỖI: {exc}")

            wb.save(BATH_IMAGE_XLSX)

    finally:
        try:
            wb.save(BATH_IMAGE_XLSX)
        except Exception:
            pass
        session.close()
        driver.quit()


if __name__ == "__main__":
    main()