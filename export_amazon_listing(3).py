import shutil
import json
from pathlib import Path
from openpyxl import load_workbook

# ================= CONFIG =================
BASE_DIR = Path(__file__).resolve().parent
STATUS_FILE = BASE_DIR / "mockup_status.xlsx"
TEMPLATE_FILE = BASE_DIR / "SHIRT_template.xlsm"

# Amazon Listing Config
BRAND_NAME = "Generic"
ITEM_TYPE_KEYWORD = "Clothing, Shoes & Jewelry > Novelty & More > Clothing > Novelty > Men > Shirts > T-Shirts (novelty-t-shirts)"
VARIATION_THEME = "SIZE/COLOR"
SIZE_SYSTEM = "US"
SIZE_CLASS = "Alpha"
SIZES = ["S", "M", "L", "XL"]
DEFAULT_PRICE = "20.99" 
DEFAULT_QUANTITY = "10"
DEFAULT_HANDLING_TIME = "3"
COLORS = {
    "w": "White",
    "b": "Black"
}

def get_image_color(url):
    """Xác định màu sắc từ URL Cloudinary (w hoặc b)."""
    url_lower = url.lower()
    if "-w-" in url_lower: return "w"
    if "-b-" in url_lower: return "b"
    return None

def export_listing_for_folder(folder_path, ai_data):
    """Điền dữ liệu vào template và lưu file xlsm."""
    folder_name = folder_path.name
    parent_sku = f"{folder_name}-001"
    
    # 1. Phân loại ảnh theo màu
    all_urls = ai_data.get("cloudinary_links", [])
    images_by_color = {"w": [], "b": []}
    for url in all_urls:
        c = get_image_color(url)
        if c: images_by_color[c].append(url)

    # 2. Mở Template (Sheet 'Template')
    wb = load_workbook(TEMPLATE_FILE, keep_vba=True)
    if 'Template' not in wb.sheetnames:
        print(f"  ❌ Lỗi: Không tìm thấy sheet 'Template' trong {TEMPLATE_FILE.name}")
        return
    ws = wb['Template']

    # Dữ liệu từ AI
    Item_Name = ai_data.get("Item_Name", "")
    description = ai_data.get("description", "")
    bullets = ai_data.get("bullets", [])
    keywords = ai_data.get("keywords", "")
    if isinstance(keywords, list): keywords = ", ".join(keywords)
    subject = ai_data.get("subject_character", "")

    current_row = 7 # Bắt đầu từ hàng 7 theo yêu cầu

    # --- TẠO DÒNG PARENT (Hàng 7) ---
    ws.cell(row=current_row, column=1).value = parent_sku              # SKU
    ws.cell(row=current_row, column=2).value = "SHIRT"                 # Product Type
    ws.cell(row=current_row, column=4).value = "Parent"                # Parentage Level
    ws.cell(row=current_row, column=6).value = VARIATION_THEME         # Variation Theme
    ws.cell(row=current_row, column=7).value = Item_Name                   # Item Name
    ws.cell(row=current_row, column=8).value = BRAND_NAME              # Brand Name
    ws.cell(row=current_row, column=11).value = ITEM_TYPE_KEYWORD      # Item Type Keyword
    ws.cell(row=current_row, column=15).value = parent_sku      # model name = sku
    ws.cell(row=current_row, column=62).value = subject                # Subject Character
    ws.cell(row=current_row, column=144).value = subject
    
    ws.cell(row=current_row, column=203).value = DEFAULT_PRICE         # List Price
    ws.cell(row=current_row, column=228).value = DEFAULT_QUANTITY      # Quantity (US)
    ws.cell(row=current_row, column=229).value = DEFAULT_HANDLING_TIME # Handling Time (US)
    
    # Ảnh đại diện cho Parent (Trắng 1)
    if images_by_color["w"]:
        ws.cell(row=current_row, column=19).value = images_by_color["w"][0]
        for i, url in enumerate(images_by_color["w"][1:9]): # Amazon cho phép tối đa 8 ảnh phụ
                    ws.cell(row=current_row, column=20 + i).value = url    # Other Images
    current_row += 1


    # --- TẠO CÁC DÒNG CHILD (Bắt đầu từ hàng 8) ---
    for color_code, color_name in COLORS.items():
        imgs = images_by_color.get(color_code, [])
        for size in SIZES:
            child_sku = f"{parent_sku}-{color_code}-{size.lower()}"
            child_item_name = f"{Item_Name} ({color_name} - {size} )" # Item Name của Child bao gồm Màu và Size
            
            ws.cell(row=current_row, column=1).value = child_sku           # SKU
            ws.cell(row=current_row, column=2).value = "SHIRT"             # Product Type
            ws.cell(row=current_row, column=4).value = "Child"             # Parentage Level
            ws.cell(row=current_row, column=5).value = parent_sku          # Parent SKU (Ánh xạ từ Parent)
            ws.cell(row=current_row, column=6).value = VARIATION_THEME     # Variation Theme
            ws.cell(row=current_row, column=7).value = child_item_name     # Item Name (Child)
            ws.cell(row=current_row, column=8).value = BRAND_NAME          # Brand Name
            ws.cell(row=current_row, column=11).value = ITEM_TYPE_KEYWORD  # Item Type Keyword
            ws.cell(row=current_row, column=232).value = DEFAULT_PRICE     # Your Price USD (Sell on Amazon, US)
            
            # New Columns for Child
            ws.cell(row=current_row, column=203).value = DEFAULT_PRICE         # List Price
            ws.cell(row=current_row, column=228).value = DEFAULT_QUANTITY      # Quantity (US)
            ws.cell(row=current_row, column=229).value = DEFAULT_HANDLING_TIME # Handling Time (US)
            
            # Mô tả & Bullets
            ws.cell(row=current_row, column=29).value = description        # Product Description
            for i, b in enumerate(bullets[:5]):
                ws.cell(row=current_row, column=30 + i).value = b          # Bullet Points
            ws.cell(row=current_row, column=35).value = keywords           # Generic Keywords

            # Size & Color & Subject
            ws.cell(row=current_row, column=46).value = SIZE_SYSTEM        # Size System
            ws.cell(row=current_row, column=47).value = SIZE_CLASS         # Size Class
            ws.cell(row=current_row, column=48).value = size               # Size Value
            ws.cell(row=current_row, column=62).value = subject            # Subject Character
            ws.cell(row=current_row, column=183).value = subject  # Animal Theme
            ws.cell(row=current_row, column=64).value = color_name         # Color Map
            ws.cell(row=current_row, column=65).value = color_name         # Color

            # Hình ảnh theo đúng màu SKU
            if imgs:
                ws.cell(row=current_row, column=19).value = imgs[0]        # Main Image
                for i, url in enumerate(imgs[1:9]): # Amazon cho phép tối đa 8 ảnh phụ
                    ws.cell(row=current_row, column=20 + i).value = url    # Other Images

            current_row += 1

    # Lưu file .xlsm
    output_path = folder_path / f"{folder_name}_Amazon_Listing.xlsm"
    wb.save(output_path)
    print(f"  ✅ Đã xuất listing: {output_path.name}")

def main():
    if not STATUS_FILE.exists() or not TEMPLATE_FILE.exists():
        print("❌ Thiếu file mockup_status.xlsx hoặc SHIRT_template.xlsm")
        return

    wb_status = load_workbook(STATUS_FILE)
    ws_status = wb_status.active

    for row_idx in range(2, ws_status.max_row + 1):
        folder_path_str = ws_status.cell(row=row_idx, column=1).value
        json_str = ws_status.cell(row=row_idx, column=2).value
        status = ws_status.cell(row=row_idx, column=4).value # Cột 4 là status chính

        if folder_path_str and json_str and status == "done":
            folder_path = Path(folder_path_str)
            print(f"📦 Đang tạo listing cho: {folder_path.name}")
            try:
                ai_data = json.loads(json_str)
                export_listing_for_folder(folder_path, ai_data)
            except Exception as e:
                print(f"  ❌ Lỗi: {e}")

    print("\n🚀 Hoàn tất!")

if __name__ == "__main__":
    main()
