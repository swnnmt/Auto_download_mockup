import os
import json
import time
from pathlib import Path
import base64
from openpyxl import load_workbook
from groq import Groq
import cloudinary
import cloudinary.uploader
from dotenv import load_dotenv
load_dotenv()

# ================= CONFIG =================
BASE_DIR = Path(__file__).resolve().parent
STATUS_FILE = BASE_DIR / "mockup_status.xlsx"

GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME")
CLOUDINARY_API_KEY = os.environ.get("CLOUDINARY_API_KEY")
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET")

# Check thiếu key
if not all([GROQ_API_KEY, CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY, CLOUDINARY_API_SECRET]):
    raise ValueError("❌ Missing environment variables. Check your .env file")

# Cloudinary Config (Bạn cần thay đổi Cloud Name, API Key, API Secret)
cloudinary.config( 
    cloud_name = CLOUDINARY_CLOUD_NAME, 
    api_key = CLOUDINARY_API_KEY, 
    api_secret = CLOUDINARY_API_SECRET, # Bạn hãy điền API Secret thật vào đây
    secure = True
)

client = Groq(api_key=GROQ_API_KEY)

MODEL_NAME = "meta-llama/llama-4-scout-17b-16e-instruct"  # Model Vision mới của Groq (Llama 4 Scout)

VALID_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}


# ================= UTIL =================
def get_images_from_folder(folder_path: Path, count=None):
    images = []
    if not folder_path.exists():
        return images

    for file in sorted(folder_path.iterdir()):
        if file.suffix.lower() in VALID_IMAGE_EXTENSIONS:
            images.append(file)
        if count and len(images) >= count:
            break
    return images


def upload_to_cloudinary(image_paths, folder_name):
    """Upload danh sách ảnh lên Cloudinary và trả về list URL."""
    uploaded_urls = []
    print(f"  -> Bắt đầu upload {len(image_paths)} ảnh lên Cloudinary folder '{folder_name}'...")
    
    for img_path in image_paths:
        try:
            # Upload ảnh lên Cloudinary vào thư mục pod/{folder_name}
            response = cloudinary.uploader.upload(
                str(img_path),
                folder=f"pod/{folder_name}",
                use_filename=True,
                unique_filename=True
            )
            url = response.get("secure_url")
            uploaded_urls.append(url)
            print(f"     ✅ Uploaded: {img_path.name} -> {url}")
        except Exception as e:
            print(f"     ❌ Lỗi khi upload {img_path.name}: {e}")
            
    return uploaded_urls


def get_mime_type(file_path: Path):
    ext = file_path.suffix.lower()
    if ext == ".jpg":
        return "image/jpeg"
    return f"image/{ext.replace('.', '')}"


# ================= AI =================
def encode_image(image_path: Path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def analyze_mockups_with_ai(image_paths, max_retries=3):
    if not image_paths:
        return None

    prompt_text = """
You are an expert Amazon Listing Optimizer. 
Analyze the provided product mockup image and return a high-converting listing in JSON format.

Return ONLY valid JSON (no markdown):
{
  "Item_Name": "",
  "bullets": ["", "", "", "", ""],
  "description": "",
  "keywords": [],
  "subject_character": ""
}

Optimization Rules:

1. Item_Name:
- Length: MIN 190 characters, MAX 200 characters.
- MUST follow EXACT structure below (do NOT remove any part):

[Adjective/Theme] [Character/Subject] [Context/Activity] T-Shirt – “[Exact Quote on Shirt]”, Heavy Cotton, [Niche/Use Case], [Occasion/Target Audience], Casual Everyday Wear, Graphic Tee, Comfortable Fit, Trendy Statement Shirt

- The quote MUST be included exactly inside quotation marks.
- The name MUST be long, natural, readable, and close to 200 characters.
- Do NOT shorten or skip any segment in the structure.
- Expand each segment slightly to reach the required length.

2. bullets:
- Exactly 5 bullet points.
- Each bullet MUST start with a CAPITALIZED FEATURE (e.g., PREMIUM COMFORT:, PERFECT GIFT:, UNIQUE DESIGN:, LIGHTWEIGHT FABRIC:, VERSATILE STYLE:).
- Focus on benefits, emotions, and use cases (gift, daily wear, humor impact).
- Mention the quote theme where relevant.

3. description:
- Write a detailed, persuasive, and engaging product description.
- Highlight the meaning, humor, or emotional value of the quote.
- Include SEO keywords naturally (quote shirt, funny saying shirt, gift idea, etc.).
- Avoid repetition, keep it readable and conversion-focused.

4. keywords:
- Provide 10–15 SEO-optimized search terms.
- Include variations of quote shirt, saying shirt, funny quote t-shirt, gift keywords, niche audience keywords.
- Comma-separated, no duplicates.

5. subject_character:
- Identify the main theme of the shirt.
- For quote-based designs, use values like: "Quote", "Funny Quote", "Motivational Quote", "Sarcastic Quote", etc.
"""

    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": prompt_text}
            ]
        }
    ]

    # Thêm ảnh vào nội dung của user message
    for img_path in image_paths:
        base64_image = encode_image(img_path)
        messages[0]["content"].append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{get_mime_type(img_path)};base64,{base64_image}"
            }
        })

    for attempt in range(max_retries):
        try:
            completion = client.chat.completions.create(
                model=MODEL_NAME,
                messages=messages,
                temperature=0.7,
                max_tokens=1024,
                top_p=1,
                stream=False,
                response_format={"type": "json_object"}
            )

            text = completion.choices[0].message.content.strip()
            data = json.loads(text)

            if "Item_Name" in data and "bullets" in data:
                return data

        except Exception as e:
            print(f"⚠️ Lỗi AI lần {attempt+1}: {e}")
            time.sleep(2)

    return None


# ================= MAIN =================
def process_generate_information():
    if not STATUS_FILE.exists():
        print(f"❌ Không tìm thấy file: {STATUS_FILE}")
        return

    wb = load_workbook(STATUS_FILE)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        folder_path_str = ws.cell(row=row_idx, column=1).value
        # Cột 2: json, Cột 3: statusUpload, Cột 4: status
        status_upload = ws.cell(row=row_idx, column=3).value
        status = ws.cell(row=row_idx, column=4).value

        if not folder_path_str:
            continue

        if status == "done":
            continue

        folder_path = Path(folder_path_str)

        print(f"\n📂 Đang xử lý: {folder_path.name}")

        # 1. Lấy toàn bộ ảnh để upload lên Cloudinary (nếu chưa upload)
        all_images = get_images_from_folder(folder_path)
        if not all_images:
            print("  -> Không có ảnh để xử lý")
            ws.cell(row=row_idx, column=4).value = "no_images"
            wb.save(STATUS_FILE)
            continue

        # Upload toàn bộ ảnh lên Cloudinary nếu statusUpload chưa là 'done'
        if status_upload != "done":
            cloudinary_urls = upload_to_cloudinary(all_images, folder_path.name)
            if cloudinary_urls:
                # Lưu tất cả link về file linkImg.txt để làm backup
                link_file_path = folder_path / "linkImg.txt"
                with open(link_file_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(cloudinary_urls))
                
                ws.cell(row=row_idx, column=3).value = "done"
                wb.save(STATUS_FILE)
                print(f"  ✅ Đã upload toàn bộ ảnh và lưu backup vào {link_file_path.name}")
            else:
                print("  ❌ Upload Cloudinary thất bại")
                continue
        else:
            print("  ⏭️  Ảnh đã được upload trước đó, bỏ qua bước upload.")
            # Đọc lại link từ file backup nếu đã upload rồi để dùng cho các bước sau
            link_file_path = folder_path / "linkImg.txt"
            if link_file_path.exists():
                with open(link_file_path, "r", encoding="utf-8") as f:
                    cloudinary_urls = [line.strip() for line in f.readlines() if line.strip()]
            else:
                # Nếu không thấy file backup thì phải upload lại
                cloudinary_urls = upload_to_cloudinary(all_images, folder_path.name)
        # 2. Lấy 1 ảnh để gửi AI phân tích marketing (giảm dung lượng tránh lỗi 413)
        ai_images = all_images[:1]
        print(f"  -> Gửi {len(ai_images)} ảnh lên AI để phân tích marketing...")

        ai_data = analyze_mockups_with_ai(ai_images)

        if ai_data:
            # Trộn link Cloudinary vào JSON kết quả
            ai_data["cloudinary_links"] = cloudinary_urls
            
            output_json = folder_path / "info.json"

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump(ai_data, f, ensure_ascii=False, indent=4)

            ws.cell(row=row_idx, column=2).value = json.dumps(ai_data, ensure_ascii=False)
            ws.cell(row=row_idx, column=4).value = "done"
            wb.save(STATUS_FILE)

            print("  ✅ Hoàn tất phân tích marketing")
        else:
            ws.cell(row=row_idx, column=4).value = "fail"
            wb.save(STATUS_FILE)
            print("  ❌ AI phân tích thất bại")

        time.sleep(2)

    print("\n🚀 Hoàn tất toàn bộ!")


# ================= RUN =================
if __name__ == "__main__":
    process_generate_information()